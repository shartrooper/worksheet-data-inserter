import datetime,logging,re
from openpyxl.styles import Font, Alignment, Border, Side
import electronbloodTestsInserterv2 as ebt

logging.basicConfig(
    level=logging.DEBUG, format="%(asctime)s -  %(levelname)s -  %(message)s"
)

class StyleUtilities:

    def strIntoUsefulRegex(self,testname):
        charList = ["\.", "\(", "\)"]
        for char in charList:
            charRegex = re.compile(rf"{char}")
            testname = charRegex.sub(char, testname)
        return testname

    def getTotalLength(self,glossary):
        count = 0
        for category in glossary:
            count += len(glossary[category])
        return count

    def setHeaderCells(self,coordinate, ws, name, line):
        nameCell = ws.cell(
            row=ws[coordinate].row, column=ws[coordinate].column + 1, value=name
        )
        ws.merge_cells(
            start_row=ws[coordinate].row,
            start_column=ws[coordinate].column + 1,
            end_row=ws[coordinate].row,
            end_column=8,
        )
        nameCell.font = Font(underline=line, name="Arial", size=12)
        nameCell.alignment = Alignment(horizontal="left", vertical="justify")

    def reassignStyles(self,ws,styledCell):
        ws[styledCell].font = Font(bold=True, name="Arial", size=12)
        ws[styledCell].alignment = Alignment(horizontal="center", vertical="center")

    def drawBottomBorder(self,ws,targetRow, maxCols):
        borderLen = 8
        # Initialized borderlen and determine closer multiple to draw stripe along row
        while maxCols > borderLen:
            borderLen *= 2
        medium = Side(border_style="medium", color="FF000000")
        for j in range(1, borderLen + 1):
            ws.cell(row=targetRow, column=j).border = Border(bottom=medium)

    def cleanBottomBorder(self,ws,startingCol):
        for row in ws.iter_rows(min_row=1, min_col=startingCol):
            for cell in row:
                removed = Side(border_style=None)
                cell.border = Border(bottom=removed)
    
    def applythinBorders(self,cell):
        topStyle=cell.border.top.style
        bottomStyle=cell.border.bottom.style
        if topStyle == "medium" and bottomStyle == "medium":
            styledborder = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),
            top=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='medium',color='FF000000'))
        elif topStyle == "medium" and not bottomStyle:
            styledborder = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),
            top=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        elif not topStyle and bottomStyle == "medium":
            styledborder = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),
            top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='medium',color='FF000000'))
        else:
            styledborder = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),
            top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        cell.border=styledborder

class GetHeaderContent:

    def __init__(self,report,title):
        self.__currentReport=report
        self.__title=title
        self.__patientName=''
        self.__patientRUT=''
        self.__reportDate=None
        self.__wsDate=None
        self.__wsTime=None
        self.__dateRE = re.compile(r"Fecha Recepci[oó]n\s?:?\s?(([0-3]?[0-9])[\/-]([0-1]?[0-9])[\/-]([1-2][0-9]{3}))\s(\d{2}:\d{2})",re.IGNORECASE)
        self.__patientRE = re.compile(r"paciente[\s]*[:]?[\s]*([a-záíúéó]+\s+[a-záíúéó]+\s+[a-záíúéó]+\s+[a-záíúéó]+)",re.IGNORECASE)
        self.__searchAndSetHeaderParams()        
    
    def __searchAndSetHeaderParams(self):
        dateRegex = self.__dateRE
        patientRegex =self.__patientRE
        rutRegex= re.compile("(\d+\.\d{3}\.\d{3})(?=-\d+)")
        try:
            for page in self.__currentReport:
                if dateRegex.search(page) and patientRegex.search(page) and rutRegex.search(page):
                    self.__reportDate = dateRegex.search(page)
                    self.__patientName = patientRegex.search(page)
                    self.__patientRUT = rutRegex.search(page)
                    #logging.debug('The patient name : '+self.__patientName.group(1)+' The report date is : '+self.__reportDate.group(0)+'his Chilean RUT is: '+self.__patientRUT.group(1))
                    break
                else:
                    raise Exception("Date or Patient's credentials not found")
            self.__wsDate = datetime.date(int(self.__reportDate.group(4)), int(self.__reportDate.group(3)), int(self.__reportDate.group(2))).strftime("%d/%m/%y")
            self.__wsTime = self.__reportDate.group(5)
        except Exception as err:
            print("An exception happened: "+str(err))
            ebt.WriteLog(str(err))

    def getWsDate(self):
        return self.__wsDate

    def getWsTime(self):
        return self.__wsTime

    def getHeaderFormat(self):
        return {"Título": self.__title,
            "Nombre": self.__patientName.group(1),
            "RUT":self.__patientRUT.group(1),
            "Fecha": None,
            "Hora": None}
    
    def getReport(self):
        return self.__currentReport
        
    def getRowPosition(self):
        return len(self.getHeaderFormat())+1

class InsertDataInWorkSheet(GetHeaderContent):

    def __init__(self,currentWorksheet,report,title,glossary):
        super().__init__(report,title)
        self.__ws=currentWorksheet
        self.__glos=glossary
        self.__dateCoordinates=""
    
    # Format list for headers and date/time
    def insertAndFormatHeaderData(self):
        formatDataDic=self.getHeaderFormat()
        currentGlossary=self.__glos
        ws=self.__ws
        sl=StyleUtilities()
        rowPos=self.getRowPosition()
        if ws.title == "new sheet":
            for i, header in enumerate(formatDataDic, start=1):
                headerCoordinate = "A" + str(i)
                if header != "Fecha" and header != "Hora":
                    underline = "none"
                    if header == "Nombre":
                        underline = "single"
                    sl.setHeaderCells(headerCoordinate, ws, formatDataDic[header], underline)
                    ws[headerCoordinate].value = header
                    continue
                ws[headerCoordinate].value = header

            for category in currentGlossary:
                sl.drawBottomBorder(ws,rowPos - 1, ws.max_column)
                # logging.debug(row.border.bottom.style)
                for j, testname in enumerate(currentGlossary[category], start=rowPos):

                    for key in testname.keys():
                        currentCell = ws.cell(row=j, column=1, value=key)
                    # logging.debug(str(currentCell.row)+' with value '+currentCell.value)
                    rowPos += 1
            col = ws.column_dimensions["A"]
            col.width = 24
            for cell in ws["A"]:
                cell.font = Font(name="Arial", size=12, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.print_title_rows = "1:2"
            ws.title = "RESULTADOS"
        elif ws['B3'].value != formatDataDic['RUT']:
            print("RUT from loaded worksheet doesn't match the one's PDF!")
            ebt.WriteLog("Patient's RUT mismatch")
        else:
            #Insert new testname and/or category bottom stripe as a WorkSheet row
            dictLen = sl.getTotalLength(currentGlossary) + len(formatDataDic)
            totalRows = ws.max_row
            if dictLen > totalRows:
                # Clean borders before run cell insertion
                sl.cleanBottomBorder(ws,1)
                # Get current column counts with nonEmpty date values
                nonEmptyCols=0
                for value in ws.iter_cols(
                        min_row=ws["A4"].row,
                        max_row=ws["A4"].row,
                        min_col=1,
                        values_only=True,
                    ):
                        if value[0]:
                            nonEmptyCols += 1
                            continue
                        break
                for category in currentGlossary:
                    if ws.cell(row=rowPos-1, column=1).border.bottom.style != "medium":
                        sl.drawBottomBorder(ws,rowPos-1,nonEmptyCols)
                    for j, testname in enumerate(currentGlossary[category], start=rowPos):
                        
                        for key in testname.keys():
                            if ws.cell(row=j, column=1).value != key:
                                ws.insert_rows(j)
                                ws.cell(row=j, column=1, value=key)
                                sl.reassignStyles(ws,ws.cell(row=j,column=1).coordinate)
                        rowPos += 1
    # Add date value to a column
    def insertAndFormatDates(self):
        formatDataDic=self.getHeaderFormat()
        ws=self.__ws
        wsDate=self.getWsDate()
        wsTime=self.getWsTime()
        sl=StyleUtilities()
        for col in ws.iter_cols(min_col=2, max_col=999999):
            if self.__dateCoordinates:
                break
            for cell in col:
                dateCell = cell.coordinate[0] + str(len(formatDataDic) - 1)
                timeCell = cell.coordinate[0] + str(len(formatDataDic))
                if not ws[dateCell].value:
                    ws[dateCell].value = wsDate
                    ws[timeCell].value = wsTime
                    self.__dateCoordinates = ws[dateCell].coordinate
                    sl.cleanBottomBorder(ws,2)
                    nonEmpyCols = 0
                    for value in ws.iter_cols(
                        min_row=ws[dateCell].row,
                        max_row=ws[dateCell].row,
                        min_col=1,
                        values_only=True,
                    ):
                        if value[0]:
                            nonEmpyCols += 1
                            continue
                        break
                    for row in ws.iter_rows(min_row=1, min_col=1, max_col=1):
                        for cell in row:
                            if ws.cell(row=cell.row, column=1).border.bottom.style == "medium":
                                sl.drawBottomBorder(ws,cell.row, nonEmpyCols)
                elif ws[dateCell].value == wsDate and ws[timeCell].value == wsTime:
                    self.__dateCoordinates = ws[dateCell].coordinate
                    sl.cleanBottomBorder(ws,2)
                    nonEmpyCols = 0
                    for value in ws.iter_cols(
                        min_row=ws[dateCell].row,
                        max_row=ws[dateCell].row,
                        min_col=1,
                        values_only=True,
                    ):
                        if value[0]:
                            nonEmpyCols += 1
                            continue
                        break
                    for row in ws.iter_rows(min_row=1, min_col=1, max_col=1):
                        for cell in row:
                            if ws.cell(row=cell.row, column=1).border.bottom.style == "medium":
                                sl.drawBottomBorder(ws,cell.row, nonEmpyCols)
                elif ws[dateCell].value != wsDate or ws[timeCell].value != wsTime:
                    wsDateParams = wsDate.split("/")
                    colDateParams = ws[dateCell].value.split("/")
                    colTimeCell = ws[timeCell].value
                    d1 = datetime.datetime(
                        int(wsDateParams[2]),
                        int(wsDateParams[1]),
                        int(wsDateParams[0]),
                        int(wsTime[:2]),
                        int(wsTime[3:5]),
                    )
                    d2 = datetime.datetime(
                        int(colDateParams[2]),
                        int(colDateParams[1]),
                        int(colDateParams[0]),
                        int(colTimeCell[:2]),
                        int(colTimeCell[3:5]),
                    )
                    if d1 < d2:
                        ws.insert_cols(cell.column, amount=1)
                        ws.cell(row=ws[dateCell].row, column=cell.column - 1, value=wsDate)
                        self.__dateCoordinates = ws.cell(
                            row=ws[dateCell].row, column=cell.column - 1
                        ).coordinate
                        newTimeCoordinate = self.__dateCoordinates[0] + str(len(formatDataDic))
                        ws[newTimeCoordinate].value = wsTime
                        sl.setHeaderCells(ws["A1"].coordinate, ws, formatDataDic["Título"], "none")
                        sl.setHeaderCells(
                            ws["A2"].coordinate, ws, formatDataDic["Nombre"], "single"
                        )
                        sl.setHeaderCells(
                            ws["A3"].coordinate, ws, formatDataDic["RUT"], "none"
                        )
                    sl.cleanBottomBorder(ws,2)
                    nonEmpyCols = 0
                    for value in ws.iter_cols(
                        min_row=ws[dateCell].row,
                        max_row=ws[dateCell].row,
                        min_col=1,
                        values_only=True,
                    ):
                        if value[0]:
                            nonEmpyCols += 1
                            continue
                        break
                    for row in ws.iter_rows(min_row=1, min_col=1, max_col=1):
                        for cell in row:
                            if ws.cell(row=cell.row, column=1).border.bottom.style == "medium":
                                sl.drawBottomBorder(ws,cell.row, nonEmpyCols)
                break
    # Add tests results to a column
    def insertTestResultData(self):
        currentGlossary=self.__glos
        sl=StyleUtilities()
        ws=self.__ws
        stream=self.getReport()
        for category in currentGlossary:
            for testname in currentGlossary[category]:
                for row in ws.rows:
                    for cell in row:
                        for label,keyword in testname.items(): 
                            if label == cell.value:
                                for page in stream:
                                    reTestname = sl.strIntoUsefulRegex(keyword)
                                    testRegex = re.compile(
                                        rf"{reTestname}[\s]*[:]?[\s]*([-*]?\d+\.?\d*)",
                                        re.IGNORECASE,
                                    )
                                    if testRegex.search(page):
                                        testResult = testRegex.search(page)
                                        ws[
                                            self.__dateCoordinates[0] + str(cell.row)
                                        ].value = testResult.group(1)
                                        break
                                break
    # Apply the thin borders on cells
    def AddThinCellBorderStyle(self):
        sl=StyleUtilities()
        ws=self.__ws
        for row in ws.rows:
            for cell in row:
                sl.applythinBorders(cell)
