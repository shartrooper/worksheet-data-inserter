# bloodTestsInserter.pyv2 get keywords from test reports with a specific format and inserts results in a .xlsx
"""
-   takes either a .txt or .pdf, a glossary of keywords and output filenames from command line
-   e.g. bloodTestsInserter.py <test report filename> <glossary> <output worksheet>
-   If a worksheet isn't loaded, creates a new one from the glossary.
-   Only inserts test results that exists in glossary
"""
import sys, PyPDF2, logging, re, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

logging.basicConfig(
    level=logging.DEBUG, format="%(asctime)s -  %(levelname)s -  %(message)s"
)


def openFiles(filenames):
    try:
        fileExtRegex = re.compile(r"txt|pdf")
        found = fileExtRegex.search(filenames[1])
        txtStream = []
        if found:
            if found.group(0) == "txt":
                txtFile = open(filenames[1])
                content = txtFile.read()
                txtStream.append(content)
            elif found.group(0) == "pdf":
                pdfFileObj = open(filenames[1], "rb")
                pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
                for numPage in range(0, pdfReader.getNumPages()):
                    pageObj = pdfReader.getPage(numPage)
                    txtStream.append(pageObj.extractText())
                    # logging.debug(txtStream[numPage])
                pdfFileObj.close()
        else:
            raise Exception("Not a valid file ext")
        dictFile = open(filenames[2], "r")
        dictStrings = dictFile.readlines()
        dictio = {}
        category = "undefined category"
        for i in range(0, len(dictStrings)):
            # Asterisk denote Category title.
            categoryRegex = re.compile(r"[\wáíúéó\s]+(?=\*)", re.IGNORECASE)
            categ = categoryRegex.search(dictStrings[i])
            if categ:
                category = dictStrings[i].strip()
                dictio[category] = []
                continue
            dictio[category].append(dictStrings[i].strip())
            # logging.debug(dictio[category])
        if len(filenames) > 3:
            wb = load_workbook(filenames[3])
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "new sheet"
        dictFile.close()
    except Exception as err:
        print("An exception happened: " + str(err))
        sys.exit(1)
    return [txtStream, dictio, wb]


def strIntoUsefulRegex(testname):
    charList = ["\.", "\(", "\)"]
    for char in charList:
        charRegex = re.compile(rf"{char}")
        testname = charRegex.sub(char, testname)
    return testname


def getTotalLength(glossary):
    count = 0
    for category in glossary:
        count += len(glossary[category])
    return count


def setHeaderCells(coordinate, ws, name, line):
    nameCell = ws.cell(
        row=ws[coordinate].row, column=ws[coordinate].column + 1, value=name
    )
    ws.merge_cells(
        start_row=ws[coordinate].row,
        start_column=ws[coordinate].column + 1,
        end_row=ws[coordinate].row,
        end_column=8,
    )
    nameCell.font = Font(underline=line, name="Arial", size=12)
    nameCell.alignment = Alignment(horizontal="left", vertical="justify")


def reassignStyles(styledCell):
    ws[styledCell].font = Font(bold=True, name="Arial", size=12)
    ws[styledCell].alignment = Alignment(horizontal="center", vertical="center")


def drawBottomBorder(targetRow, maxCols):
    borderLen = 8
    # Initialized borderlen and determine closer multiple to draw stripe along row
    while maxCols > borderLen:
        borderLen *= 2
    medium = Side(border_style="medium", color="FF000000")
    for j in range(1, borderLen + 1):
        ws.cell(row=targetRow, column=j).border = Border(bottom=medium)


def cleanBottomBorder(startingCol):
    for row in ws.iter_rows(min_row=1, min_col=startingCol):
        for cell in row:
            removed = Side(border_style=None)
            cell.border = Border(bottom=removed)


if len(sys.argv) >= 3:
    # Get arguments from command line.
    stream, dictionary, wb = openFiles(sys.argv)
else:
    print(
        "Please, introduce at least existing filenames for test report and dictionary in command lines"
    )
    sys.exit(1)

# If apply, create workBook with testnames, date a patient's name rows
dateRegex = re.compile(
    r"Fecha Recepci[oó]n\s?:?\s?(([0-3]?[1-9])[\/-]([0-1]?[1-9])[\/-]([1-2][0-9]{3}))\s(\d{2}:\d{2})",
    re.IGNORECASE,
)

pacientRegex = re.compile(
    r"paciente[\s]*[:]?[\s]*([a-záíúéó]+\s+[a-záíúéó]+\s+[a-záíúéó]+\s+[a-záíúéó]+)",
    re.IGNORECASE,
)

for page in stream:
    if dateRegex.search(page) and pacientRegex.search(page):
        reportDate = dateRegex.search(page)
        patientName = pacientRegex.search(page)
        # logging.debug('The patient name : '+patientName.group(1)+' The report date is : '+reportDate.group())
        break
    else:
        raise Exception("Date or Pacient's name not found")

wsDate = datetime.date(
    int(reportDate.group(4)), int(reportDate.group(3)), int(reportDate.group(2))
).strftime("%d/%m/%y")
wsTime = reportDate.group(5)
ws = wb.active
dateCoordinates = ""

# Format list for headers ad date/time

formatDataDic = {
    "Título": "Resumen exámenes de laboratorio UPC HGF",
    "Nombre": patientName.group(1),
    "Fecha": None,
    "Hora": None,
}
rowPos = len(formatDataDic) + 1

if ws.title == "new sheet":
    for i, header in enumerate(formatDataDic, start=1):
        headerCoordinate = "A" + str(i)
        if header == "Título" or header == "Nombre":
            underline = "none"
            if header == "Nombre":
                underline = "single"
            setHeaderCells(headerCoordinate, ws, formatDataDic[header], underline)
            ws[headerCoordinate].value = header
            continue
        ws[headerCoordinate].value = header

    for category in dictionary:
        drawBottomBorder(rowPos - 1, ws.max_column)
        # logging.debug(row.border.bottom.style)
        for j, testname in enumerate(dictionary[category], start=rowPos):
            currentCell = ws.cell(row=j, column=1, value=testname)
            # logging.debug(str(currentCell.row)+' with value '+currentCell.value)
            rowPos += 1
    col = ws.column_dimensions["A"]
    col.width = 24
    for cell in ws["A"]:
        cell.font = Font(name="Arial", size=12, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.print_title_rows = "1:2"
    ws.title = "RESULTADOS"
else:
    #Insert new testname and/or category bottom stripe as a WorkSheet row
    dictLen = getTotalLength(dictionary) + len(formatDataDic)
    totalRows = ws.max_row
    if dictLen > totalRows:
        # Clean borders before run cell insertion
        cleanBottomBorder(1)
        # Get current column counts with nonEmpty values
        nonEmptyCols=0
        for value in ws.iter_cols(
                min_row=ws["A3"].row,
                max_row=ws["A3"].row,
                min_col=1,
                values_only=True,
            ):
                if value[0]:
                    nonEmptyCols += 1
                    continue
                break
        for category in dictionary:
            if ws.cell(row=rowPos-1, column=1).border.bottom.style != "medium":
                drawBottomBorder(rowPos-1,nonEmptyCols)
            for j, testname in enumerate(dictionary[category], start=rowPos):
                if ws.cell(row=j, column=1).value != testname:
                    ws.insert_rows(j)
                    ws.cell(row=j, column=1, value=testname)
                    reassignStyles(ws.cell(row=j,column=1).coordinate)
                rowPos += 1

# Add date value to a column
for col in ws.iter_cols(min_col=2, max_col=999999):
    if dateCoordinates:
        break
    for cell in col:
        dateCell = cell.coordinate[0] + str(len(formatDataDic) - 1)
        timeCell = cell.coordinate[0] + str(len(formatDataDic))
        if not ws[dateCell].value:
            ws[dateCell].value = wsDate
            ws[timeCell].value = wsTime
            dateCoordinates = ws[dateCell].coordinate
            cleanBottomBorder(2)
            nonEmpyCols = 0
            for value in ws.iter_cols(
                min_row=ws[dateCell].row,
                max_row=ws[dateCell].row,
                min_col=1,
                values_only=True,
            ):
                if value[0]:
                    nonEmpyCols += 1
                    continue
                break
            for row in ws.iter_rows(min_row=1, min_col=1, max_col=1):
                for cell in row:
                    if ws.cell(row=cell.row, column=1).border.bottom.style == "medium":
                        drawBottomBorder(cell.row, nonEmpyCols)
        elif ws[dateCell].value == wsDate and ws[timeCell].value == wsTime:
            dateCoordinates = ws[dateCell].coordinate
            cleanBottomBorder(2)
            nonEmpyCols = 0
            for value in ws.iter_cols(
                min_row=ws[dateCell].row,
                max_row=ws[dateCell].row,
                min_col=1,
                values_only=True,
            ):
                if value[0]:
                    nonEmpyCols += 1
                    continue
                break
            for row in ws.iter_rows(min_row=1, min_col=1, max_col=1):
                for cell in row:
                    if ws.cell(row=cell.row, column=1).border.bottom.style == "medium":
                        drawBottomBorder(cell.row, nonEmpyCols)
        elif ws[dateCell].value != wsDate or ws[timeCell].value != wsTime:
            wsDateParams = wsDate.split("/")
            colDateParams = ws[dateCell].value.split("/")
            colTimeCell = ws[timeCell].value
            d1 = datetime.datetime(
                int(wsDateParams[2]),
                int(wsDateParams[1]),
                int(wsDateParams[0]),
                int(wsTime[:2]),
                int(wsTime[3:5]),
            )
            d2 = datetime.datetime(
                int(colDateParams[2]),
                int(colDateParams[1]),
                int(colDateParams[0]),
                int(colTimeCell[:2]),
                int(colTimeCell[3:5]),
            )
            if d1 < d2:
                ws.insert_cols(cell.column, amount=1)
                ws.cell(row=ws[dateCell].row, column=cell.column - 1, value=wsDate)
                dateCoordinates = ws.cell(
                    row=ws[dateCell].row, column=cell.column - 1
                ).coordinate
                newTimeCoordinate = dateCoordinates[0] + str(len(formatDataDic))
                ws[newTimeCoordinate].value = wsTime
                setHeaderCells(ws["A1"].coordinate, ws, formatDataDic["Título"], "none")
                setHeaderCells(
                    ws["A2"].coordinate, ws, formatDataDic["Nombre"], "single"
                )
            cleanBottomBorder(2)
            nonEmpyCols = 0
            for value in ws.iter_cols(
                min_row=ws[dateCell].row,
                max_row=ws[dateCell].row,
                min_col=1,
                values_only=True,
            ):
                if value[0]:
                    nonEmpyCols += 1
                    continue
                break
            for row in ws.iter_rows(min_row=1, min_col=1, max_col=1):
                for cell in row:
                    if ws.cell(row=cell.row, column=1).border.bottom.style == "medium":
                        drawBottomBorder(cell.row, nonEmpyCols)
        break

# Match and insert existing testname results in textStream

def applythinBorders(cell):
    topStyle=cell.border.top.style
    bottomStyle=cell.border.bottom.style
    if topStyle == "medium" and bottomStyle == "medium":
        styledborder = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),
        top=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='medium',color='FF000000'))
    elif topStyle == "medium" and not bottomStyle:
        styledborder = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),
        top=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    elif not topStyle and bottomStyle == "medium":
        styledborder = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),
        top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='medium',color='FF000000'))
    else:
        styledborder = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),
        top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    cell.border=styledborder

for category in dictionary:
    for testname in dictionary[category]:
        for row in ws.rows:
            for cell in row:
                if testname == cell.value:
                    for page in stream:
                        reTestname = strIntoUsefulRegex(testname)
                        testRegex = re.compile(
                            rf"{reTestname}[\s]*[:]?[\s]*([-*]?\d+\.?\d*)",
                            re.IGNORECASE,
                        )
                        if testRegex.search(page):
                            testResult = testRegex.search(page)
                            ws[
                                dateCoordinates[0] + str(cell.row)
                            ].value = testResult.group(1)
                            break
                    break

for row in ws.rows:
    for cell in row:
        applythinBorders(cell)

# Save WorkBook on a file

wb.save("report" + ".xlsx")
print("Workbook updated or saved in local dir!")
