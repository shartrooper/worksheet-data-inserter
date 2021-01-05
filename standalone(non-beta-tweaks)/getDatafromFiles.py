# Refactor of bloodTestsInserter-v2.py using classes
import sys, PyPDF2, os, logging, re
from pikepdf import Pdf
from openpyxl import Workbook, load_workbook

logging.basicConfig(
    level=logging.DEBUG, format="%(asctime)s -  %(levelname)s -  %(message)s"
)


class GetDataStreamCollection:

    def __init__(self, filename):
        self.__filename = filename
        self.__txtStream = []

    def __appendFileToCollection(self):
        try:
            fileExtRegex = re.compile(r"txt|pdf")
            found = fileExtRegex.search(self.__filename)
            if found:
                if found.group(0) == "txt":
                    txtFile = open(self.__filename)
                    content = txtFile.read()
                    self.__txtStream.append(content)
                elif found.group(0) == "pdf":
                    pdfFileObj = open(self.__filename, "rb")
                    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
                    for numPage in range(0, pdfReader.getNumPages()):
                        pageObj = pdfReader.getPage(numPage)
                        self.__txtStream.append(pageObj.extractText())
                    # logging.debug(self.__txtStream[numPage])
                    pdfFileObj.close()
            else:
                raise Exception("Not a valid file ext")
        except Exception as err:
            if str(err) == "Could not find xref table at specified location":
                self.__retrieveDataFromCorruptedPDF()
                return None
            print("An exception happened: " + str(err))
            WriteLog(str(err))
            return 'error'
    
    def __retrieveDataFromCorruptedPDF(self):
        tempName='temp.pdf'
        with Pdf.open(self.__filename) as fixedPdf:
            fixedPdf.save(tempName)
            self.__filename= tempName
            self.__appendFileToCollection()
        path = os.path.join(os.getcwd(), tempName)
        os.unlink(path)
        
    def getCollection(self):
        if self.__appendFileToCollection() == 'error':
            return None
        return self.__txtStream


class GetGlossary:
    def __init__(self, glossary):
        self.__glossary = glossary
        self.__dict = {}

    def __appendStringToCollection(self):
        try:
            # intoList=self.__glossary.split('\n')
            intoList = self.__glossary.readlines()
            if intoList:
                category = "undefined category"
                categoryRegex = re.compile(r"[\wáíúéó\s]+(?=\*)", re.IGNORECASE)
                labelAndKeyRegex = re.compile(r"([\w\d\.\s%ñáíúéó\(\)]+)\s*:\s*([\w\d\.\s%ñáíúéó\(\)]+)", re.IGNORECASE)
                for i in range(0, len(intoList)):
                    categ = categoryRegex.search(intoList[i])
                    if categ:
                        category = intoList[i].strip()
                        self.__dict[category] = []
                        continue
                    labelAndKey = labelAndKeyRegex.search(intoList[i])
                    if labelAndKey:
                        self.__dict[category].append({labelAndKey.group(1).strip(): labelAndKey.group(2).strip()})
                if not len(self.__dict):
                    raise Exception("Empty Dictionary")
            else:
                raise Exception("Empty glossary")

        except Exception as err:
            print("An exception happened: " + str(err))
            WriteLog(str(err))
            sys.exit(1)

    def getCollection(self):
        self.__appendStringToCollection()
        # for category in self.__dict:
        #    logging.debug('category : '+category)
        #    logging.debug(self.__dict[category])
        return self.__dict


class LoadOrCreateWorkBook:
    def __init__(self, filename):
        self.__filename = filename

    def getWorkBook(self):
        if self.__filename:
            return load_workbook(self.__filename)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "new sheet"
            return wb


class WriteLog:

    def __logHeader(self):
        return f'===REPORTE DE ACCESO A LA APLICACIÓN===\n Total de Accesos: {self.__userCounter}\n\n'

    def __writeStatus(self):
        return f'Status acceso {self.__userCounter} : {self.__message}'

    def __init__(self, message):
        self.__message = message
        self.__logPath = "init with log file path"
        self.__userCounter = 0
        self.__createOrUpdateLog()

    def __createOrUpdateLog(self):
        # Path
        path = os.path.join(os.getcwd(), 'user_log.txt')
        if os.path.isfile(path):
            logContent = ''
            with open("user_log.txt", "r") as log:
                counterRegex = re.compile(r"total de accesos:\s?(\d+)", re.IGNORECASE)
                logContent = log.read()
                if not logContent:
                    raise Exception("Empty file")
                getCurrentCount = counterRegex.search(logContent)
                self.__userCounter = int(getCurrentCount.group(1)) + 1
                logContent = counterRegex.sub(f'Total de Accesos: {self.__userCounter}', logContent)
            # remove unnecessary log file
            os.unlink(path)
            with open("user_log.txt", "w") as log:
                log.write(logContent)
                log.seek(0, 2)
                log.write('\n' + self.__writeStatus())
            #sys.exit(1)
        else:
            with open("user_log.txt", "w") as log:
                self.__userCounter += 1
                log.write(self.__logHeader())
                log.write(self.__writeStatus())
            #sys.exit(1)