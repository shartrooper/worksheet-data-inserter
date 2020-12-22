import datetime,logging,re
from openpyxl.styles import Font, Alignment, Border, Side
import getDatafromFiles as ebt

logging.basicConfig(
    level=logging.DEBUG, format="%(asctime)s -  %(levelname)s -  %(message)s"
)

class StyleUtilities:

    def strIntoUsefulRegex(self,testname):
        charList = ["\.", "\(", "\)"]
        for char in charList:
            charRegex = re.compile(rf"{char}")
            testname = charRegex.sub(char, testname)
        isTransferrinaRegex=re.compile(r"^transferrina$",re.IGNORECASE)
        if isTransferrinaRegex.search(testname):
            return "(?<!% saturacion de )transferrina"
        return testname
    
    def strIntoNum(self,label,num):
        #remove blank space and asterisks between thousand unit (1000) and cent unit (100).
        num=re.sub('\s','',num)
        num=re.sub('\*','',num)
        num=float(num)
        #round into two decimals max, apply pH and Troponina exception.
        if num.is_integer():
            return int(num)
        elif label == 'pH':
            return float(num)
        elif label == 'Troponina':
            if num < 0.047:
                return '<0.047'
        return round(num,2)
    
    def getTotalLength(self,glossary):
        count = 0
        for category in glossary:
            count += len(glossary[category])
        return count

    def setHeaderCells(self,coordinate, ws, name, line):
        nameCell = ws.cell(
            row=ws[coordinate].row, column=ws[coordinate].column + 1, value=name
        )
        ws.merge_cells(
            start_row=ws[coordinate].row,
            start_column=ws[coordinate].column + 1,
            end_row=ws[coordinate].row,
            end_column=12,
        )
        nameCell.font = Font(underline=line, name="Arial", size=8)
        nameCell.alignment = Alignment(horizontal="center", vertical="justify")

    def reassignStyles(self,ws,styledCell,inBold):
        ws[styledCell].font = Font(bold=inBold, name="Arial", size=8)
        ws[styledCell].alignment = Alignment(horizontal="center", vertical="center")

    def drawBottomBorder(self,ws,targetRow, maxCols):
        borderLen = 12
        # Initialized borderlen and determine closer multiple to draw stripe along row
        while maxCols > borderLen:
            borderLen *= 2
        medium = Side(border_style="medium", color="FF000000")
        for j in range(1, borderLen + 1):
            ws.cell(row=targetRow, column=j).border = Border(bottom=medium)

    def cleanBottomBorder(self,ws,startingCol):
        for row in ws.iter_rows(min_row=1, min_col=startingCol):
            for cell in row:
                removed = Side(border_style=None)
                cell.border = Border(bottom=removed)
    
    def applythinBorders(self,cell):
        topStyle=cell.border.top.style
        bottomStyle=cell.border.bottom.style
        if topStyle == "medium" and bottomStyle == "medium":
            styledborder = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),
            top=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='medium',color='FF000000'))
        elif topStyle == "medium" and not bottomStyle:
            styledborder = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),
            top=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        elif not topStyle and bottomStyle == "medium":
            styledborder = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),
            top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='medium',color='FF000000'))
        else:
            styledborder = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),
            top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        cell.border=styledborder

class GetHeaderContent:

    def __init__(self,report,title):
        self.__currentReport=report
        self.__title=title
        self.__patientName=''
        self.__patientRUT=''
        self.__reportDate=None
        self.__wsDate=None
        self.__wsTime=None
        self.__dateRE = re.compile(r"Fecha Recepci[oó]n\s?:?\s?(([0-3]?[0-9])[\/-]([0-1]?[0-9])[\/-]([1-2][0-9]{3}))\s(\d{2}:\d{2})",re.IGNORECASE)
        self.__patientRE = re.compile(r"paciente[\s]*[:]?[\s]*([a-zñáíúéó\.]+ +[a-zñáíúéó\.]+ ?[a-zñáíúéó\.]* ?[a-zñáíúéó\.]* ?[a-zñáíúéó\.]*)",re.IGNORECASE)
        self.isError= False
        self.__searchAndSetHeaderParams()

    def __searchAndSetHeaderParams(self):
        dateRegex = self.__dateRE
        patientRegex =self.__patientRE
        #rutRegex= re.compile("(\d+\.\d{3}\.\d{3})(?=-[\d\w]+)")
        rutRegex= re.compile("(\d+\.\d{3}\.\d{3}-[\d\w])")
        try:
            for page in self.__currentReport:
                if dateRegex.search(page) and patientRegex.search(page) and rutRegex.search(page):
                    self.__reportDate = dateRegex.search(page)
                    self.__patientName = patientRegex.search(page)
                    self.__patientRUT = rutRegex.search(page)
                    #logging.debug('The patient name : '+self.__patientName.group(1)+' The report date is : '+self.__reportDate.group(0)+'his Chilean RUT is: '+self.__patientRUT.group(1))
                    break
                else:
                    raise Exception("Date or Patient's credentials not found")
            self.__wsDate = datetime.date(int(self.__reportDate.group(4)), int(self.__reportDate.group(3)), int(self.__reportDate.group(2))).strftime("%d/%m/%y")
            self.__wsTime = self.__reportDate.group(5)
            return None
        except Exception as err:
            print("An exception happened: "+str(err))
            ebt.WriteLog(str(err))
            self.isError= True

    def getWsDate(self):
        return self.__wsDate

    def getWsTime(self):
        return self.__wsTime

    def getHeaderFormat(self):
        return {"Título": self.__title,
            "Nombre": self.__patientName.group(1).strip(),
            "RUT":self.__patientRUT.group(1),
            "Fecha": None,
            "Hora": None}
    
    def getReport(self):
        return self.__currentReport
        
    def getRowPosition(self):
        return len(self.getHeaderFormat())+1
        
    def getErrorFlag(self):
        return self.isError

class InsertDataInWorkSheet(GetHeaderContent):

    def __init__(self,currentWorksheet,report,title,glossary):
        super().__init__(report,title)
        self.__ws=currentWorksheet
        self.__glos=glossary
        self.__dateCoordinates=""
        self._isCapped=False
    # Get current Date's cell coordinates
    def getCurrentDateCoordinates(self):
        return self.__dateCoordinates
    # Format list for headers and date/time
    def insertAndFormatHeaderData(self):
        formatDataDic=self.getHeaderFormat()
        currentGlossary=self.__glos
        ws=self.__ws
        sl=StyleUtilities()
        rowPos=self.getRowPosition()
        if ws.title == "new sheet":
            for i, header in enumerate(formatDataDic, start=1):
                headerCoordinate = "A" + str(i)
                if header != "Fecha" and header != "Hora":
                    underline = "none"
                    if header == "Nombre":
                        underline = "single"
                    sl.setHeaderCells(headerCoordinate, ws, formatDataDic[header], underline)
                    ws[headerCoordinate].value = header
                    continue
                ws[headerCoordinate].value = header

            for category in currentGlossary:
                sl.drawBottomBorder(ws,rowPos - 1, ws.max_column)
                # logging.debug(row.border.bottom.style)
                for j, testname in enumerate(currentGlossary[category], start=rowPos):

                    for key in testname.keys():
                        currentCell = ws.cell(row=j, column=1, value=key)
                    # logging.debug(str(currentCell.row)+' with value '+currentCell.value)
                    rowPos += 1
            col = ws.column_dimensions["A"]
            col.width = 15
            for cell in ws["A"]:
                cell.font = Font(name="Arial", size=8, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.print_title_rows = "1:3"
            ws.title = "RESULTADOS"
        elif ws['B3'].value != formatDataDic['RUT']:
            print("RUT from loaded worksheet doesn't match the one's PDF!")
            ebt.WriteLog("Patient's RUT mismatch")
            self.isError= True
        else:
            #Insert new testname and/or category bottom stripe as a WorkSheet row
            dictLen = sl.getTotalLength(currentGlossary) + len(formatDataDic)
            totalRows = ws.max_row
            if dictLen > totalRows:
                # Clean borders before run cell insertion
                sl.cleanBottomBorder(ws,1)
                # Get current column counts with nonEmpty date values
                nonEmptyCols=0
                for value in ws.iter_cols(
                        min_row=ws["A4"].row,
                        max_row=ws["A4"].row,
                        min_col=1,
                        values_only=True,
                    ):
                        if value[0]:
                            nonEmptyCols += 1
                            continue
                        break
                for category in currentGlossary:
                    if ws.cell(row=rowPos-1, column=1).border.bottom.style != "medium":
                        sl.drawBottomBorder(ws,rowPos-1,nonEmptyCols)
                    for j, testname in enumerate(currentGlossary[category], start=rowPos):
                        
                        for key in testname.keys():
                            if ws.cell(row=j, column=1).value != key:
                                ws.insert_rows(j)
                                ws.cell(row=j, column=1, value=key)
                                sl.reassignStyles(ws,ws.cell(row=j,column=1).coordinate,True)
                        rowPos += 1
    # Add date value to a column
    def insertAndFormatDates(self):
        formatDataDic=self.getHeaderFormat()
        ws=self.__ws
        wsDate=self.getWsDate()
        wsTime=self.getWsTime()
        sl=StyleUtilities()
        #Verify there are non-empty columns within format column's range
        for cells in ws['B4':'M4']:
            for cell in cells:
                if not cell.value:
                   break
                elif cell.coordinate == 'M4':
                    print('There are not more empty columns anymore!')
                    ebt.WriteLog('Max Columns format reached!')
                    self._isCapped=True
        #iterate through columns and insert date's data
        for col in ws.iter_cols(min_col=2, max_col=12):
            if self.__dateCoordinates:
                break
            for cell in col:
                dateCell = cell.coordinate[0] + str(len(formatDataDic) - 1)
                timeCell = cell.coordinate[0] + str(len(formatDataDic))
                sl.reassignStyles(ws,dateCell,False)
                sl.reassignStyles(ws,timeCell,False)
                if not ws[dateCell].value:
                    ws[dateCell].value = wsDate
                    ws[timeCell].value = wsTime
                    self.__dateCoordinates = ws[dateCell].coordinate
                    sl.cleanBottomBorder(ws,2)
                    nonEmpyCols = 0
                    for value in ws.iter_cols(
                        min_row=ws[dateCell].row,
                        max_row=ws[dateCell].row,
                        min_col=1,
                        values_only=True,
                    ):
                        if value[0]:
                            nonEmpyCols += 1
                            continue
                        break
                    for row in ws.iter_rows(min_row=1, min_col=1, max_col=1):
                        for cell in row:
                            if ws.cell(row=cell.row, column=1).border.bottom.style == "medium":
                                sl.drawBottomBorder(ws,cell.row, nonEmpyCols)
                    break
                elif ws[dateCell].value == wsDate and ws[timeCell].value == wsTime:
                    self.__dateCoordinates = ws[dateCell].coordinate
                    sl.cleanBottomBorder(ws,2)
                    nonEmpyCols = 0
                    for value in ws.iter_cols(
                        min_row=ws[dateCell].row,
                        max_row=ws[dateCell].row,
                        min_col=1,
                        values_only=True,
                    ):
                        if value[0]:
                            nonEmpyCols += 1
                            continue
                        break
                    for row in ws.iter_rows(min_row=1, min_col=1, max_col=1):
                        for cell in row:
                            if ws.cell(row=cell.row, column=1).border.bottom.style == "medium":
                                sl.drawBottomBorder(ws,cell.row, nonEmpyCols)
                    break
                elif cell.coordinate[0] == 'L':
                    ebt.WriteLog('Max Columns format reached!')
                    break
                elif ws[dateCell].value != wsDate or ws[timeCell].value != wsTime:
                    wsDateParams = wsDate.split("/")
                    colDateParams = ws[dateCell].value.split("/")
                    colTimeCell = ws[timeCell].value
                    d1 = datetime.datetime(
                        int(wsDateParams[2]),
                        int(wsDateParams[1]),
                        int(wsDateParams[0]),
                        int(wsTime[:2]),
                        int(wsTime[3:5]),
                    )
                    d2 = datetime.datetime(
                        int(colDateParams[2]),
                        int(colDateParams[1]),
                        int(colDateParams[0]),
                        int(colTimeCell[:2]),
                        int(colTimeCell[3:5]),
                    )
                    if d1 < d2:
                        ws.insert_cols(cell.column, amount=1)
                        ws.cell(row=ws[dateCell].row, column=cell.column - 1, value=wsDate)
                        self.__dateCoordinates = ws.cell(
                            row=ws[dateCell].row, column=cell.column - 1
                        ).coordinate
                        newTimeCoordinate = self.__dateCoordinates[0] + str(len(formatDataDic))
                        ws[newTimeCoordinate].value = wsTime
                        sl.setHeaderCells(ws["A1"].coordinate, ws, formatDataDic["Título"], "none")
                        sl.setHeaderCells(
                            ws["A2"].coordinate, ws, formatDataDic["Nombre"], "single"
                        )
                        sl.setHeaderCells(
                            ws["A3"].coordinate, ws, formatDataDic["RUT"], "none"
                        )
                        sl.reassignStyles(ws,self.__dateCoordinates,False)
                        sl.reassignStyles(ws,newTimeCoordinate,False)
                    sl.cleanBottomBorder(ws,2)
                    nonEmpyCols = 0
                    for value in ws.iter_cols(
                        min_row=ws[dateCell].row,
                        max_row=ws[dateCell].row,
                        min_col=1,
                        values_only=True,
                    ):
                        if value[0]:
                            nonEmpyCols += 1
                            continue
                        break
                    for row in ws.iter_rows(min_row=1, min_col=1, max_col=1):
                        for cell in row:
                            if ws.cell(row=cell.row, column=1).border.bottom.style == "medium":
                                sl.drawBottomBorder(ws,cell.row, nonEmpyCols)
                    break
    # Add tests results to a column
    def insertTestResultData(self):
        currentGlossary=self.__glos
        sl=StyleUtilities()
        ws=self.__ws
        stream=self.getReport()
        for category in currentGlossary:
            for testname in currentGlossary[category]:
                for row in ws.rows:
                    for cell in row:
                        for label,keyword in testname.items():
                            if label == cell.value:
                                for page in stream:
                                    reTestname = sl.strIntoUsefulRegex(keyword)
                                    testRegex = re.compile(
                                        rf"{reTestname}[\s]*[:]?[\s]*([-*]?\d*\s*[-*]?\d+\.?\d*)",
                                        re.IGNORECASE,
                                    )
                                    if testRegex.search(page):
                                        testResult = testRegex.search(page)
                                        intoNum = sl.strIntoNum(label,testResult.group(1))
                                        ws[
                                            self.__dateCoordinates[0] + str(cell.row)
                                        ].value = intoNum
                                        sl.reassignStyles(ws,ws[self.__dateCoordinates[0] + str(cell.row)].coordinate,False)
                                        break
                                break
    # Apply the thin borders on cells
    def addThinCellBorder(self):
        sl=StyleUtilities()
        ws=self.__ws
        for row in ws.rows:
            for cell in row:
                sl.applythinBorders(cell)
    #change results column dimensions
    def setDataColumnDimensions(self):
        ws=self.__ws
        for row in ws.rows:
            for cell in row:
                ws.row_dimensions[cell.row].height=10
        for col in ws.iter_cols(min_col=2):
            for cell in col:
                currentColumn=cell.coordinate[0]
                ws.column_dimensions[currentColumn].width=6
    #Delete all columns out of format's range
    def removeColSurplus(self):
        self.__ws.delete_cols(13,30)
    def isColCapReached(self):
        return self._isCapped

class AdjustCalciumValue:
    
    def __init__(self,currentWorksheet,coordinates):
        self.__ws=currentWorksheet
        self.__dateCoordinates=coordinates
        self.__albuminaValue=0
        self.__calciumValue=0
        self.__albuminaCoordinates=''
        self.__calciumCoordinates=''
        self.__getValues()
        self.__setAdjustedValue()
        
    def __correctCalciumValue(self):
        return self.__calciumValue+(4-self.__albuminaValue)*0.8
    
    def __getValues(self):
        ws=self.__ws
        dateCoord=self.__dateCoordinates
        #if 'Albumina' in ws.iter_rows(min_row=1, min_col=1, max_col=1, values_only= True): and 'Calcio (corregido)' in ws.iter_rows(min_row=1, min_col=1, max_col=1, values_only= True):
        for row in ws.iter_rows(min_row=1, min_col=1, max_col=1):
            for cell in row:
                if cell.value == 'Albumina':
                    for col in ws.iter_cols(min_row=cell.row,max_row=cell.row,min_col=2,max_col=ws[dateCoord].column):
                        for cell in col:
                            if cell.value:
                                self.__albuminaCoordinates=cell.coordinate
                                self.__albuminaValue=cell.value
                elif cell.value == 'Calcio (corregido)':
                    calciumCell=ws.cell(row=cell.row,column=ws[dateCoord].column)
                    self.__calciumCoordinates=calciumCell.coordinate
                    self.__calciumValue=calciumCell.value
    
    def __setAdjustedValue(self):
        if self.__calciumValue and self.__albuminaValue:
            self.__ws[self.__calciumCoordinates].value=self.__correctCalciumValue()

class CreateRecycleWorkSheet:
    
    def __init__(self,workBook,currentWorksheet,startingRow):
        self.__ws=currentWorksheet
        if "RECICLAJE" in workBook.sheetnames:
            workBook.remove_sheet(workBook["RECICLAJE"])
        self.__wsr=workBook.copy_worksheet(currentWorksheet)
        self.__wsr.title="RECICLAJE"
        self.__startRow=startingRow
        self.__cleanColumnsAndRows()
        self.__insertColumnInRecycleWs()

    def __cleanColumnsAndRows(self):
        wsr=self.__wsr
        nullBorder= Border(left=Side(border_style=None,color='FF000000'),
        right=Side(border_style=None,color='FF000000'),
        top=Side(border_style=None,color='FF000000'),
        bottom=Side(border_style=None, color='FF000000'))
        for col in wsr.columns:
            for cell in col:
                cell.value=''
                cell.border=nullBorder
        
    def __insertColumnInRecycleWs(self):
        ws=self.__ws
        wsr=self.__wsr
        sl=StyleUtilities()
        # Iterate through result columns and search for last
        for i,col in enumerate(ws.iter_cols(min_col=2,min_row=self.__startRow),2):
            for j,cell in enumerate(col,self.__startRow):
                # get cell from next col to compare values with
                nextCell=ws.cell(row=self.__startRow,column=i+1)
                if not nextCell.value:
                    wsrCell=wsr.cell(row=j,column=i)
                    wsrCell.value=cell.value
                    wsrCell.font = Font(bold=False, name="Arial", size=8)
                    wsrCell.alignment = Alignment(horizontal="center", vertical="top")
                    #sl.reassignStyles(wsr,wsrCell.coordinate,False)
                    continue
                break